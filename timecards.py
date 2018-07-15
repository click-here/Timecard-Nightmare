from openpyxl import Workbook
import random
import numpy as np
import pandas as pd
import re
import string
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side

names_df = pd.read_csv(r'https://raw.githubusercontent.com/click-here/Baby-Name-Popularity/master/top1000.csv')

col_headers = ['Projects', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Total']
names = ["Emma","Olivia","Ava","Isabella","Sophia","Mia","Charlotte","Amelia","Evelyn","Abigail","Harper","Emily","Elizabeth","Avery","Sofia","Ella","Madison","Scarlett","Victoria","Aria","Liam","Noah","William","James","Logan","Benjamin","Mason","Elijah","Oliver","Jacob","Lucas","Michael","Alexander","Ethan","Daniel","Matthew","Aiden","Henry","Joseph","Jackson"]


def generate_weekly_hours(median,days_per_week):
    return np.round(np.random.normal(median,2,days_per_week)*4)/4

def rng2tuple(rng): # adapted from https://stackoverflow.com/a/12640614
    col,row = re.split('(\D*)',rng)[1:]
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num, int(row)

def get_rng(col,row):
    return get_column_letter(col) + str(row)

def generate_proj_hours(projects):
    shift_length = 8
    proj_cnt = len(projects)
    project_split = shift_length / proj_cnt
    return np.round(np.random.normal(project_split, 1, proj_cnt) * 4) / 4

def calc_top_right(rng):
    c,r = rng2tuple(rng)
    c = c + len(col_headers[1:])
    return get_rng(c,r)


    
def build_timecard_table(cnt):
    wb = Workbook()
   

    projects = ['Redesign Mobile UI',
                'Write unit tests',
                'Raspberry PI Project',
                'Power BI Dashboard',
                'Fuzzy Search Implementation',
                'Dynamics Plugin Development',
                'SCOTUS Opinions Data Science',
                'Build CSS Piano Animation',
                'Honeypot Project']

    dest_filename = 'timecards\\timecard_%s.xlsx'%cnt

    ws = wb.active
    ws.title = "Timecard"

    ws['F2'] = 'Employee ID'
    ws['G2'] = '0000%s'%(cnt)
    ws['F3'] = 'Employee'
    ws['G3'] = names_df.sample(1)['name'].iloc[0]
    ws['F4'] = 'Week Ending'
    ws['G4'] = '7/21/2018'
    ws['F5'] = 'Manager'
    ws['G5'] = names_df.sample(1)['name'].iloc[0]

    # set the top left of the hours table
    top_left_cell = 'C7'
    top_right_cell = calc_top_right(top_left_cell)

    project_count = random.randint(1, 4)
    timecard_projects = random.sample(projects, project_count)

    #bold the header row
    header_row = '%s:%s'%(top_left_cell,top_right_cell)
    for row in ws[header_row]:
        for cell in row:
            cell.number_format = '0'
            cell.font = Font(bold=True,color='1061B6')
    
    for col in range(len(col_headers)):
        header = col_headers[col]
        c,r = rng2tuple(top_left_cell)
        active_col = c + col
        ws.cell(column=active_col, row=r, value= header)

        task_hours = generate_proj_hours(timecard_projects)
        for proj in range(len(timecard_projects)):
            active_row = r + proj + 1
            if header == 'Projects':
                project = timecard_projects[proj]
                ws.cell(column=active_col, row=active_row, value=project)
            elif header == 'Total':
                rbound = get_rng(active_col-1,active_row)
                lbound = get_rng(active_col-5, active_row)
                ws.cell(column=active_col, row=active_row, value='=sum(%s:%s)'%(lbound,rbound))
            else:
                ws.cell(column=active_col, row=active_row, value=task_hours[proj])
    c, r = rng2tuple(top_left_cell)

    # write total for week formula

    tbound = get_rng(active_col, active_row)
    bbound = get_rng(active_col, active_row - proj)
    ws.cell(column=active_col, row=active_row + 1, value='=sum(%s:%s)'%(tbound,bbound))

    wb.save(filename=dest_filename)

for i in range(100):
    build_timecard_table(i)
    


